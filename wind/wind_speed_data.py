from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

MEAN_WIND_SPEED_SHEET = "Mean Wind Speed Data"
MAX_WIND_SPEED_SHEET = "Max Wind Speed Data"


def store_data(data_sets):
    filename = '/tmp/ForthRoadBridgeWindSpeedReport.xlsx'
    book = open_or_create(filename)

    for sheet in book:
        data = None
        if sheet.title == MEAN_WIND_SPEED_SHEET:
            data = data_sets["mean"]
        elif sheet.title == MAX_WIND_SPEED_SHEET:
            data = data_sets["max"]

        if data:
            start_row = first_empty_row(sheet, 1)
            end_row = start_row + len(data)
            for i in range(start_row, end_row):
                sheet.cell(row = i, column = 1).value = data[i - start_row]["timestamp"]
                sheet.cell(row = i, column = 2).value = data[i - start_row]["speed"]
            sheet.column_dimensions['A'].width = 25

    book = _create_chart(book)

    book.save(filename)

    return filename

def _create_chart(book):
    c1 = LineChart()
    c1.title = "Mean Wind Speed"
    c1.style = 13
    c1.y_axis.title = 'Speed'
    c1.x_axis.title = 'Time'

    mean_sheet = book.get_sheet_by_name(MEAN_WIND_SPEED_SHEET)
    max_sheet = book.get_sheet_by_name(MAX_WIND_SPEED_SHEET)
    max_row = first_empty_row(mean_sheet, 1) - 1
    mean_data = Reference(mean_sheet, min_col=2, min_row=1, max_col=2, max_row=max_row)
    max_data = Reference(max_sheet, min_col=2, min_row=1, max_col=2, max_row=max_row)
    dates = Reference(mean_sheet, min_col=1, min_row=2, max_row=max_row)

    c1.add_data(mean_data, titles_from_data=True)
    c1.add_data(max_data, titles_from_data=True)
    c1.y_axis.scaling.max = 100
    c1.set_categories(dates)
    
    s1 = c1.series[0]
    s2 = c1.series[1]
    s1.graphicalProperties.line.solidFill = "FF0000"
    s2.graphicalProperties.line.solidFill = "0000FF"
    
    ws = book.create_sheet()
    ws.title = "Charts"
    ws.add_chart(c1, "A1")

    return book

def first_empty_row(sheet, col):
    i = 1
    while sheet.cell(row = i, column = col).value:
        i +=1
    return i

def open_or_create(file_name):
    book = Workbook()
    book.active.title = MEAN_WIND_SPEED_SHEET
    book.active['A1'].value = "Time Stamp"
    book.active['B1'].value = "Average Wind Speed (mph)"
    
    sheet2 = book.create_sheet()
    sheet2.title = MAX_WIND_SPEED_SHEET
    sheet2['A1'].value = "Time Stamp"
    sheet2['B1'].value = "Max Wind Speed (mph)"

    return book